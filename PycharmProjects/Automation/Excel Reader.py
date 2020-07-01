# Create a program to read thousands of spread-sheets
import openpyxl as xl
from openpyxl.chart import BarChart3D, Reference, PieChart


def sales_func(filename):
    # Load the file and the sheet tab
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    # Create a new price column based on the original price
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    # Bar chart
    values = Reference(sheet, min_row=1, max_row=sheet.max_row, min_col=3, max_col=4)
    transaction_id = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=1, max_col=1)
    chart = BarChart3D()
    chart.title = '3D Bar Chart'
    chart.x_axis.title = "Transaction ID"
    chart.y_axis.title = "Adjusted Price ($)"
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(transaction_id)
    sheet.add_chart(chart, 'E2')

    # Pie chart
    quantities = Reference(sheet, min_row=1, max_row=5, min_col=13, max_col=13)
    labels = Reference(sheet, min_row=2, max_row=5, min_col=12, max_col=12)
    chart2 = PieChart()
    chart2.title = 'Pie Chart'
    chart2.add_data(quantities, titles_from_data=True)
    chart2.set_categories(labels)
    sheet.add_chart(chart2, 'N2')

    wb.save(filename)


sales_func('transactions.xlsx')